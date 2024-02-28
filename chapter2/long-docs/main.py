# -*- coding: utf-8 -*-
import os
from openpyxl import load_workbook
from diff_docs import diff

file_path = os.path.join(os.path.dirname(__file__), '../data/doc_urls.xlsx')

wb = load_workbook(file_path)
sheet = wb.active
for row in sheet.iter_rows(values_only=True, min_row=2):
    url_cn = row[0]
    url_intl = row[0].replace('help.aliyun.com', 'www.alibabacloud.com/help')

    print('------')
    print(f'中国站文档：{url_cn}')
    print(f'国际站文档：{url_intl}')
    result = diff(url_cn, url_intl)
    print(result)

